# === NON-PROGRAMMER GUIDE ===
# Purpose: Implements the office xlsx parser part of the application runtime.
# What to read first: Start at the top-level function/class definitions and follow calls downward.
# Inputs: Configuration values, command arguments, or data files used by this module.
# Outputs: Returned values, written files, logs, or UI updates produced by this module.
# Safety notes: Update small sections at a time and run relevant tests after edits.
# ============================
# ============================================================================
# HybridRAG -- XLSX Parser (src/parsers/office_xlsx_parser.py)
# ============================================================================
#
# WHAT THIS FILE DOES (plain English):
#   Reads Excel (.xlsx) spreadsheet files and converts them into
#   searchable text. Each sheet becomes a labeled section, and each
#   row becomes a pipe-delimited line (like: "Part No | Description | Qty").
#
# WHY THIS MATTERS:
#   Parts lists, test results, configuration tables, and inventory
#   spreadsheets contain structured data that people need to search.
#   "What part number is the antenna?" should find the row in the
#   equipment spreadsheet even though it's not a PDF or Word doc.
#
# HOW IT WORKS:
#   1. Open the .xlsx file using openpyxl in read-only mode (low RAM)
#   2. Loop through every sheet (workbook tab)
#   3. Tag each sheet with [SHEET] SheetName for traceability
#   4. For each row, convert all cell values to strings
#   5. Join cells with " | " pipe delimiters
#   6. Skip entirely empty rows (all cells blank)
#   7. Return combined text + details (sheet count, row count)
#
# LIMITATIONS:
#   This is not "perfect Excel understanding" -- it does not interpret
#   formulas, charts, or conditional formatting. But it makes the
#   TEXT CONTENT of every cell searchable, which covers 90%+ of use cases.
#
# INTERNET ACCESS: NONE
# ============================================================================

from __future__ import annotations

from pathlib import Path
from typing import Tuple, Dict, Any
from datetime import datetime, date


def _cell_to_text(value: Any) -> str:
    """Normalize Excel cell values into stable searchable text."""
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        # ISO style is compact, sortable, and unambiguous for retrieval.
        return value.isoformat(sep=" ", timespec="seconds") if isinstance(value, datetime) else value.isoformat()
    return str(value).strip()


class XlsxParser:
    """
    Extract text from Excel .xlsx spreadsheets.

    Opens the workbook read-only via openpyxl, walks every sheet and
    every cell, and concatenates all non-empty cell values separated
    by tabs (columns) and newlines (rows). Sheet names are included
    as section headers so search results can reference which sheet
    the data came from.
    """

    def parse(self, file_path: str) -> str:
        text, _ = self.parse_with_details(file_path)
        return text

    def parse_with_details(self, file_path: str) -> Tuple[str, Dict[str, Any]]:
        path = Path(file_path)
        details: Dict[str, Any] = {"file": str(path), "parser": "XlsxParser"}

        try:
            import openpyxl
        except Exception as e:
            details["error"] = f"IMPORT_ERROR: {type(e).__name__}: {e}"
            return "", details

        try:
            # data_only=True reads the computed VALUES of formulas (not the
            # formula text). read_only=True prevents loading the entire file
            # into memory at once (important for huge spreadsheets).
            wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)

            parts = []
            sheets = 0
            rows_emitted = 0

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheets += 1
                parts.append(f"[SHEET] {sheet_name}")
                header = None
                header_row_idx = None

                for ridx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    vals = [_cell_to_text(v) for v in row]
                    if all(x == "" for x in vals):
                        continue

                    # First non-empty row is treated as header if it looks tabular.
                    if header is None and any(x != "" for x in vals[1:]):
                        header = [v if v else f"col_{i+1}" for i, v in enumerate(vals)]
                        header_row_idx = ridx
                        parts.append(f"[HEADER row={ridx}] " + " | ".join(header))
                        continue

                    rows_emitted += 1
                    # Always include raw row for exact matching.
                    parts.append(f"[ROW {ridx}] " + " | ".join(vals))

                    # Header-aware representation improves retrieval for
                    # "part number for X" style queries.
                    if header:
                        keyed = []
                        for i, cell in enumerate(vals):
                            if cell == "":
                                continue
                            col_name = header[i] if i < len(header) else f"col_{i+1}"
                            keyed.append(f"{col_name}: {cell}")
                        if keyed:
                            parts.append(f"[ROW_KV {ridx}] " + " ; ".join(keyed))

                if header_row_idx is not None:
                    parts.append(f"[SHEET_SUMMARY] header_row={header_row_idx}")

            full = "\n".join(parts).strip()
            details["total_len"] = len(full)
            details["sheets"] = sheets
            details["rows_emitted"] = rows_emitted

            wb.close()
            return full, details
        except Exception as e:
            details["error"] = f"RUNTIME_ERROR: {type(e).__name__}: {e}"
            return "", details
