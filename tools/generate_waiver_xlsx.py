"""
Generate waiver_cheat_sheet_v5.xlsx -- comprehensive software approval workbook.
FILE: tools/generate_waiver_xlsx.py

Run: python tools/generate_waiver_xlsx.py
Output: docs/05_security/waiver_cheat_sheet_v5.xlsx

CHANGELOG:
  2026-03-01: v5 -- Full rebuild from waiver_reference_sheet.md v5c.
              Added justification notes, CVE status, transitive deps,
              BLUE recommendations with business cases, RETIRED section.
"""
import os
import sys
from datetime import date

# Add project root so openpyxl import works from venv
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, PROJECT_ROOT)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# STYLE DEFINITIONS
# ---------------------------------------------------------------------------
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
DATA_FONT = Font(name="Calibri", size=10)
TITLE_FONT = Font(name="Calibri", bold=True, size=14)
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="top")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Colors
GREEN_FILL = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="DAA520", end_color="DAA520", fill_type="solid")
BLUE_FILL = PatternFill(start_color="4169E1", end_color="4169E1", fill_type="solid")
RED_FILL = PatternFill(start_color="CC3333", end_color="CC3333", fill_type="solid")
GRAY_FILL = PatternFill(start_color="708090", end_color="708090", fill_type="solid")
LIGHT_GREEN = PatternFill(start_color="F0FFF0", end_color="F0FFF0", fill_type="solid")
LIGHT_YELLOW = PatternFill(start_color="FFFFF0", end_color="FFFFF0", fill_type="solid")
LIGHT_BLUE = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
LIGHT_RED = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")
LIGHT_GRAY = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")


def _apply_header(ws, row, cols, fill):
    """Apply header styling to a row."""
    for c, val in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _apply_data_row(ws, row, vals, fill=None):
    """Write a data row with styling."""
    for c, val in enumerate(vals, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = DATA_FONT
        cell.alignment = WRAP_ALIGN
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill


def _set_col_widths(ws, widths):
    """Set column widths."""
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# DATA
# ---------------------------------------------------------------------------

GREEN_HEADERS = ["#", "Package", "Version", "License", "Publisher / Origin",
                 "Category", "Purpose", "Justification Notes", "CVE Status"]

GREEN_DATA = [
    # Core Runtime
    [1, "Python", "3.11.9", "PSF-2.0", "Python.org / USA", "Core Runtime",
     "Runtime environment",
     "Ships with OS or standalone installer. Foundation for all Python packages.",
     "No known CVEs"],
    [2, "pip", "26.0.1", "MIT", "PyPA / USA", "Core Runtime",
     "Package installer (ships with Python)",
     "Auto-installed with Python. Required to install any other package.",
     "No known CVEs"],
    [3, "setuptools", "65.5.0", "MIT", "PyPA / USA", "Core Runtime",
     "Build tools (ships with Python)",
     "Auto-installed with Python. Required by pip for package building.",
     "No known CVEs"],
    # AI / Model APIs
    [4, "numpy", "1.26.4", "BSD-3", "NumFOCUS / USA", "AI / Numerical",
     "Vector math, embedding storage, similarity computation",
     "Universal foundation for ALL Python numerical/scientific work. If any AI/ML tool is approved, numpy is implicitly required. Pinned to 1.x (2.x has breaking changes).",
     "No known CVEs"],
    [5, "tiktoken", "0.8.0", "MIT", "OpenAI / USA", "AI / Tokenization",
     "Token counting (offline, no network)",
     "Counts tokens for cost estimation and context window management. Runs entirely offline with no network calls. Published by OpenAI, same vendor as Azure OpenAI.",
     "No known CVEs"],
    # HTTP and Networking
    [6, "httpx", "0.28.1", "BSD-3", "Encode / UK", "HTTP / Networking",
     "Async HTTP client for embedder + openai SDK",
     "Modern HTTP client required by openai SDK and used by our Ollama embedder. Five Eyes ally (UK). Zero telemetry.",
     "No known CVEs"],
    [7, "requests", "2.32.5", "Apache 2.0", "PSF / USA", "HTTP / Networking",
     "HTTP client (pre-installed with pip)",
     "Ships with pip. Industry standard HTTP library used by millions of Python projects.",
     "No known CVEs"],
    [8, "urllib3", "2.6.3", "MIT", "PSF / USA", "HTTP / Networking",
     "HTTP internals (pre-installed with pip)",
     "Low-level HTTP library. Dependency of requests and pip itself. Already present in any Python installation.",
     "No known CVEs"],
    [9, "cryptography", "44.0.2", "Apache 2.0", "PyCA / USA", "HTTP / Networking",
     "AES-256 encryption, SSL/TLS",
     "Pre-installed with pip. Handles all encryption operations. Published by Python Cryptographic Authority (USA).",
     "No known CVEs"],
    [10, "certifi", "2026.1.4", "MPL-2.0", "Kenneth Reitz / USA", "HTTP / Networking",
     "SSL certificate bundle (requests dep)",
     "Mozilla CA certificate bundle for HTTPS verification. Dependency of requests.",
     "No known CVEs"],
    # Data Validation
    [11, "pydantic", "2.11.1", "MIT", "Pydantic / USA", "Data Validation",
     "Data validation and serialization (openai SDK dep)",
     "Required dependency of openai SDK. Also used for config validation throughout the project. MIT license, USA publisher.",
     "No known CVEs"],
    [12, "pyyaml", "6.0.2", "MIT", "Kirill Simonov / USA", "Data Validation",
     "YAML config file parsing",
     "Reads configuration files. Already on approved software list. No network activity.",
     "No known CVEs"],
    # Document Parsers
    [13, "pdfplumber", "0.11.9", "MIT", "Jeremy Singer-Vine / USA", "Document Parsers",
     "Primary PDF text extraction",
     "Extracts text from PDF documents for indexing. MIT license. Used by major news organizations for data journalism.",
     "No known CVEs"],
    [14, "pdfminer.six", "20251230", "MIT", "Community / USA", "Document Parsers",
     "PDF parsing engine (pdfplumber dep)",
     "Low-level PDF parsing library. Required dependency of pdfplumber.",
     "No known CVEs"],
    [15, "pypdf", "6.6.2", "BSD-3", "Community / USA", "Document Parsers",
     "PDF metadata and page counting",
     "Lightweight PDF library for metadata extraction and page counting. No external dependencies.",
     "No known CVEs"],
    [16, "pypdfium2", "5.3.0", "Apache 2.0", "Google / USA", "Document Parsers",
     "PDF rendering (same engine as Chrome)",
     "Uses the same Google PDFium engine that powers Chrome's built-in PDF viewer. If Chrome/Edge is approved, this uses identical code.",
     "No known CVEs"],
    [17, "pdf2image", "1.17.0", "MIT", "Community / USA", "Document Parsers",
     "PDF to image conversion for OCR pipeline",
     "Converts PDF pages to images for OCR processing. MIT license, no network activity.",
     "No known CVEs"],
    [18, "pytesseract", "0.3.13", "Apache 2.0", "Community / USA", "Document Parsers",
     "OCR bridge for scanned PDFs",
     "Python wrapper for Tesseract OCR engine. Enables reading scanned/image-based PDFs. Apache 2.0 license.",
     "No known CVEs"],
    [19, "python-docx", "1.2.0", "MIT", "Community / USA", "Document Parsers",
     "Word .docx reader",
     "Reads Microsoft Word files. Only READS Office XML format -- same format Microsoft Office uses. If Office is approved, this reads the same files.",
     "No known CVEs"],
    [20, "python-pptx", "1.0.2", "MIT", "Community / USA", "Document Parsers",
     "PowerPoint .pptx reader",
     "Reads Microsoft PowerPoint files. Read-only access to Office Open XML format.",
     "No known CVEs"],
    [21, "openpyxl", "3.1.5", "MIT", "Community / USA", "Document Parsers",
     "Excel .xlsx reader",
     "Reads/writes Microsoft Excel files. Same Office Open XML format. Used for this waiver spreadsheet itself.",
     "No known CVEs"],
    [22, "xlsxwriter", "3.2.9", "BSD-2", "John McNamara / USA", "Document Parsers",
     "Excel .xlsx writer (reports, exports)",
     "Write-only Excel library for generating reports and data exports. BSD-2 license.",
     "No known CVEs"],
    [23, "lxml", "6.0.2", "BSD-3", "Community / USA", "Document Parsers",
     "XML/HTML parsing (Office XML dependency)",
     "XML parser required by python-docx, python-pptx, openpyxl. All Office file formats are XML internally.",
     "No known CVEs"],
    [24, "pillow", "12.1.0", "HPND", "PIL / USA", "Document Parsers",
     "Image processing for OCR and PDF rendering",
     "Python Imaging Library. Handles image operations for PDF rendering and OCR pipeline. HPND license (permissive, similar to BSD).",
     "No known CVEs"],
    # Web Server / API
    [25, "fastapi", "0.115.0", "MIT", "Tiangolo / USA", "Web Server / API",
     "REST API framework (localhost only)",
     "High-performance REST API framework. Binds to localhost only (127.0.0.1:8000). No external network exposure. MIT license.",
     "No known CVEs"],
    [26, "uvicorn", "0.41.0", "BSD-3", "Encode / UK", "Web Server / API",
     "ASGI server (runs FastAPI)",
     "Production ASGI server for FastAPI. Five Eyes ally (UK). Binds localhost only.",
     "No known CVEs"],
    [27, "starlette", "0.38.6", "BSD-3", "Encode / UK", "Web Server / API",
     "ASGI toolkit (FastAPI dependency)",
     "Low-level web toolkit required by FastAPI. Version 0.38.6 includes DoS fix (CVE in <0.38.2 patched).",
     "Patched in 0.38.2+"],
    # Configuration and Logging
    [28, "structlog", "24.4.0", "MIT", "Hynek Schlawack / Germany", "Logging",
     "Structured JSON logging for audit trails",
     "Generates structured JSON log entries for security audit trails. Five Eyes ally (Germany/EU). No network activity.",
     "No known CVEs"],
    [29, "rich", "13.9.4", "MIT", "Will McGugan / UK", "Logging",
     "Console formatting (display only)",
     "Terminal output formatting. Display-only, no data processing or network activity. Five Eyes ally (UK).",
     "No known CVEs"],
    [30, "tqdm", "4.67.3", "MIT", "Community / USA", "Logging",
     "Progress bars (display only)",
     "Progress bar display during indexing operations. Display-only, no data or network activity.",
     "No known CVEs"],
    [31, "regex", "2026.1.15", "Apache 2.0", "Community / USA", "Text Processing",
     "Enhanced regular expressions",
     "Extended regex library for text processing. Drop-in replacement for stdlib re with better Unicode support.",
     "No known CVEs"],
    [32, "colorama", "0.4.6", "BSD-3", "Community / USA", "Logging",
     "Console colors (Windows terminal support)",
     "Enables ANSI color codes on Windows terminals. Display-only, zero network activity.",
     "No known CVEs"],
    # Credential Storage
    [33, "keyring", "23.13.1", "MIT", "Jason R. Coombs / USA", "Credential Storage",
     "Windows Credential Manager access",
     "Stores API keys in Windows Credential Manager (OS-level encrypted vault). Never writes credentials to disk files.",
     "No known CVEs"],
    # Other Direct Dependencies
    [34, "python-multipart", "0.0.22", "Apache 2.0", "Community / USA", "Web Server / API",
     "Form data parsing (FastAPI dependency)",
     "Parses multipart form data for file uploads via API. Required by FastAPI.",
     "No known CVEs"],
    [35, "click", "8.3.1", "BSD-3", "Pallets / USA", "Web Server / API",
     "CLI toolkit (uvicorn dependency)",
     "Command-line interface library. Required by uvicorn for CLI argument parsing. Published by Pallets (Flask maintainers).",
     "No known CVEs"],
]

GREEN_TRANSITIVE_HEADERS = ["#", "Package", "Version", "License", "Pulled In By",
                            "Notes"]

GREEN_TRANSITIVE_DATA = [
    [1, "annotated-types", "0.7.0", "MIT", "pydantic", "Type annotation support"],
    [2, "anyio", "4.12.1", "MIT", "httpx", "Async I/O abstraction"],
    [3, "cffi", "2.0.0", "MIT", "cryptography", "C FFI for OpenSSL bindings"],
    [4, "chardet", "5.2.0", "LGPL-2.1", "pdfminer.six", "Character encoding detection"],
    [5, "charset-normalizer", "3.4.4", "MIT", "requests", "Encoding normalization"],
    [6, "distro", "1.9.0", "Apache 2.0", "openai", "Linux distro detection"],
    [7, "et_xmlfile", "2.0.0", "MIT", "openpyxl", "XML utility for Excel"],
    [8, "h11", "0.16.0", "MIT", "uvicorn", "HTTP/1.1 protocol parser"],
    [9, "httpcore", "1.0.9", "BSD-3", "httpx", "Low-level HTTP transport"],
    [10, "idna", "3.11", "BSD-3", "requests", "International domain names"],
    [11, "jaraco.classes", "3.4.0", "MIT", "keyring", "Class utilities"],
    [12, "jiter", "0.13.0", "MIT", "pydantic", "Fast JSON parsing (Rust)"],
    [13, "markdown-it-py", "4.0.0", "MIT", "rich", "Markdown parsing"],
    [14, "mdurl", "0.1.2", "MIT", "markdown-it-py", "URL parsing for markdown"],
    [15, "more-itertools", "10.8.0", "MIT", "keyring", "Iterator utilities"],
    [16, "packaging", "26.0", "Apache 2.0", "pytest", "Version parsing"],
    [17, "pycparser", "3.0", "BSD-3", "cffi", "C parser for FFI"],
    [18, "pydantic_core", "2.33.0", "MIT", "pydantic", "Core validation engine (Rust)"],
    [19, "Pygments", "2.19.2", "BSD-2", "rich", "Syntax highlighting"],
    [20, "pywin32-ctypes", "0.2.3", "BSD-3", "keyring (Windows)", "Windows API access"],
    [21, "sniffio", "1.3.1", "MIT", "httpx", "Async library detection"],
    [22, "typing_extensions", "4.15.0", "PSF-2.0", "pydantic", "Python typing backports"],
    [23, "typing-inspection", "0.4.2", "MIT", "pydantic", "Type introspection"],
    [24, "iniconfig", "2.3.0", "MIT", "pytest", "INI file parsing for pytest config"],
    [25, "pluggy", "1.6.0", "MIT", "pytest", "Plugin framework for pytest"],
]

YELLOW_HEADERS = ["#", "Package", "Version", "License", "Publisher / Origin",
                  "Category", "Purpose", "Justification Notes",
                  "Data Flow", "Network Activity", "CVE Status"]

YELLOW_DATA = [
    [1, "openai", "1.109.1", "MIT", "OpenAI / USA", "AI API Client",
     "API client for Azure OpenAI cloud queries",
     "Industry-standard SDK for OpenAI-compatible APIs. Recommended by Microsoft for Azure OpenAI Service. MIT licensed, zero telemetry. PINNED to v1.x (never upgrade to 2.x -- breaking API syntax changes). Bumped 1.51.2->1.109.1 for httpx 0.28 compatibility.",
     "HTTPS to single configured Azure endpoint only",
     "One outbound HTTPS request per user query (online mode only; offline mode uses zero network)",
     "No known CVEs"],
    [2, "pytest", "9.0.2", "MIT", "Holger Krekel / Germany", "Testing",
     "Test framework (410+ regression tests)",
     "Standard Python test framework used by 90%+ of Python projects. All 410+ regression tests depend on it. Development-only tool -- not deployed to production. MIT license, zero network activity. Five Eyes ally (Germany/EU).",
     "None -- runs tests locally",
     "Zero network activity",
     "No known CVEs"],
    [3, "psutil", "7.2.2", "BSD-3", "Giampaolo Rodola / USA", "System Monitoring",
     "Process/memory monitoring during index builds",
     "Monitors CPU, memory, and disk usage during document indexing to prevent resource exhaustion. BSD-3 license, USA publisher. No network activity.",
     "Reads OS process/memory stats only",
     "Zero network activity",
     "No known CVEs"],
    [4, "Ollama", "latest stable", "MIT", "Ollama Inc. / USA", "AI Runtime",
     "Run AI language models locally -- zero cloud, zero internet",
     "Enables fully offline AI-powered document search and Q&A. All models run locally with no data leaving the machine. Localhost binding only (127.0.0.1:11434). MIT license, USA company. Backbone of zero-trust offline architecture.",
     "localhost only (127.0.0.1:11434)",
     "Zero outbound connections during operation. Model downloads are one-time via CLI (ollama pull).",
     "No known CVEs"],
]

YELLOW_MODELS_HEADERS = ["#", "Model", "Size", "License", "Publisher / Origin",
                         "Purpose", "Justification Notes"]

YELLOW_MODELS_DATA = [
    [1, "nomic-embed-text", "274 MB", "Apache 2.0", "Nomic AI / USA",
     "Document embeddings (required)",
     "Converts text to 768-dim vectors for semantic search. Required for all retrieval operations. Replaces retired HuggingFace models (2.5 GB savings)."],
    [2, "phi4-mini", "2.5 GB", "MIT", "Microsoft / USA",
     "Primary Q&A model (laptop)",
     "Lightweight 3.8B parameter model for offline Q&A. MIT license, Microsoft publisher. 128K context window."],
    [3, "mistral:7b", "4.4 GB", "Apache 2.0", "Mistral AI / France",
     "Engineering alternate model",
     "7B parameter model for engineering-focused queries. Apache 2.0, France (Five Eyes+ ally). Fallback for gen profile."],
    [4, "phi4:14b-q4_K_M", "9.1 GB", "MIT", "Microsoft / USA",
     "High-accuracy workstation model (primary for 8/9 profiles)",
     "14B parameter model, quantized to 4-bit. Primary model for desktop/workstation. MIT license, Microsoft/USA. Best accuracy in approved stack."],
    [5, "gemma3:4b", "3.3 GB", "Apache 2.0", "Google / USA",
     "Fast summarization and PM fallback",
     "4B parameter model optimized for summarization. Apache 2.0, Google/USA. PM profile fallback."],
    [6, "mistral-nemo:12b", "7.1 GB", "Apache 2.0", "Mistral AI + NVIDIA",
     "Long document processing (128K context)",
     "12B parameter model with 128K token context window. Gen profile primary. Joint Mistral/NVIDIA release. Apache 2.0."],
]

BLUE_HEADERS = ["#", "Package", "Version", "License", "Publisher / Origin",
                "Purpose", "Business Case", "Dependencies",
                "Data Flow", "Timeline"]

BLUE_DATA = [
    [1, "faiss-cpu", "1.9.0", "MIT", "Meta AI Research / USA",
     "Fast vector similarity search (10-100x faster than brute-force)",
     "Current brute-force numpy search will not scale beyond 500K documents. faiss-cpu provides approximate nearest neighbor search used by every major search engine. Pure MIT license (no AI model restrictions -- this is a library, not a model). Single dependency (numpy, already approved). CPU-only version, no GPU drivers needed.",
     "numpy (already approved)",
     "In-process library, zero network activity",
     "Needed when document corpus exceeds 100K chunks"],
    [2, "lancedb", "0.29.2", "Apache 2.0", "LanceDB Inc. / USA (YC-backed)",
     "Embedded vector database replacing SQLite+memmap+FTS5 triple-store",
     "Current architecture uses three separate stores (SQLite for metadata, memmap for vectors, FTS5 for keyword search). LanceDB unifies all three into a single embedded database. 'The SQLite of vector databases.' Zero telemetry, file-based persistence, 4 MB idle memory. Apache 2.0, San Francisco USA company.",
     "pyarrow (Apache Foundation), pydantic (approved), numpy (approved)",
     "Fully embedded, serverless, zero network activity",
     "Recommended for next major version"],
    [3, "pyarrow", ">=16.0", "Apache 2.0", "Apache Software Foundation / USA",
     "Columnar data format (lancedb dependency)",
     "Apache Arrow in-memory columnar format. Required by lancedb. Published by Apache Software Foundation. Industry standard for high-performance data processing.",
     "None (self-contained)",
     "In-process library, zero network activity",
     "Apply together with lancedb"],
    [4, "vllm", "0.10.1", "Apache 2.0", "UC Berkeley / USA",
     "GPU-optimized LLM serving (batching, prefix caching, tensor parallelism)",
     "Enables serving larger models efficiently on dual RTX 3090 (48 GB VRAM). Provides batched inference, continuous batching, and tensor parallelism. Required for running 24B+ models. Conflicts with current openai==1.109.1 pin (needs openai>=1.99.1). Apache 2.0, UC Berkeley/USA.",
     "openai>=1.99.1 (upgrade required)",
     "localhost serving only",
     "Not needed until workstation hardware arrives"],
]

RED_HEADERS = ["#", "Software", "Publisher", "Country", "Category",
               "Reason for Ban", "Alternative"]

RED_DATA = [
    [1, "Qwen (all versions)", "Alibaba", "China", "AI Model",
     "NDAA restricted entity. Chinese government-affiliated corporation.",
     "phi4:14b-q4_K_M (Microsoft/USA, MIT)"],
    [2, "DeepSeek (all versions)", "DeepSeek", "China", "AI Model",
     "NDAA restricted entity. Chinese AI company.",
     "mistral-nemo:12b (Mistral/France, Apache 2.0)"],
    [3, "BGE / BGE-M3 embeddings", "BAAI", "China", "AI Model",
     "NDAA restricted entity. Beijing Academy of AI.",
     "nomic-embed-text (Nomic AI/USA, Apache 2.0)"],
    [4, "Llama (all versions)", "Meta", "USA", "AI Model",
     "Meta AUP prohibits weapons/military use -- creates regulatory conflict.",
     "phi4:14b-q4_K_M (Microsoft/USA, MIT)"],
    [5, "Milvus / pymilvus", "Zilliz", "China (Shanghai)", "Vector DB",
     "NDAA -- China origin, core engineering team in China.",
     "faiss-cpu (Meta AI Research/USA, MIT) or lancedb (LanceDB/USA, Apache 2.0)"],
    [6, "LangChain", "LangChain Inc.", "USA", "Framework",
     "200+ transitive dependencies, chronic version instability, massive attack surface.",
     "Direct API calls via openai SDK (MIT)"],
    [7, "ChromaDB", "Chroma Inc.", "USA", "Vector DB",
     "onnxruntime bloat, posthog telemetry phoning home, Windows compatibility issues.",
     "lancedb (LanceDB/USA, Apache 2.0)"],
    [8, "PyMuPDF", "Artifex", "USA", "PDF Library",
     "AGPL copyleft license -- viral, requires source disclosure of entire application.",
     "pdfplumber + pypdf (MIT/BSD)"],
    [9, "DuckDB + VSS", "DuckDB Foundation", "Netherlands", "Vector DB",
     "VSS extension is experimental -- data corruption risk on crash.",
     "lancedb or faiss-cpu"],
    [10, "Qdrant (local)", "Qdrant GmbH", "Germany", "Vector DB",
     "Local mode explicitly 'dev only'; 400 MB constant RAM overhead.",
     "lancedb (4 MB idle memory)"],
    [11, "PostgreSQL + pgvector", "Community", "International", "Vector DB",
     "Requires separate server process -- violates embedded/portable requirement.",
     "lancedb (fully embedded, serverless)"],
]

RETIRED_HEADERS = ["#", "Package", "Last Version", "License",
                   "Reason for Removal", "Replaced By", "Savings"]

RETIRED_DATA = [
    [1, "sentence-transformers", "2.7.0", "Apache 2.0",
     "HuggingFace models require AI Use Case approval workflow",
     "Ollama nomic-embed-text (274 MB)", "~800 MB"],
    [2, "torch (PyTorch)", "2.10.0", "BSD-3",
     "No longer needed without HuggingFace models",
     "Ollama handles inference", "~1.5 GB"],
    [3, "transformers", "4.57.6", "Apache 2.0",
     "No longer needed without HuggingFace models",
     "Ollama handles inference", "~100 MB"],
    [4, "tokenizers", "0.22.2", "Apache 2.0",
     "No longer needed without HuggingFace models",
     "tiktoken (OpenAI, already approved)", "~20 MB"],
    [5, "huggingface_hub", "0.36.1", "Apache 2.0",
     "HuggingFace models retired entirely",
     "Ollama model management", "~15 MB"],
    [6, "safetensors", "0.7.0", "Apache 2.0",
     "No longer needed without HuggingFace models",
     "N/A", "~5 MB"],
    [7, "scipy", "1.17.0", "BSD-3",
     "Was sentence-transformers dependency",
     "numpy (already approved) handles remaining math", "~50 MB"],
    [8, "scikit-learn", "1.8.0", "BSD-3",
     "BM25 keyword search now handled by SQLite FTS5",
     "SQLite FTS5 (built into Python stdlib)", "~30 MB"],
]


# ---------------------------------------------------------------------------
# WORKBOOK GENERATION
# ---------------------------------------------------------------------------

def create_workbook():
    wb = Workbook()

    # -----------------------------------------------------------------------
    # SHEET 1: Summary
    # -----------------------------------------------------------------------
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = "4169E1"

    ws.merge_cells("A1:E1")
    cell = ws.cell(row=1, column=1, value="Software Applications Waiver Reference -- v5")
    cell.font = TITLE_FONT

    ws.cell(row=2, column=1, value="Project:").font = SUBTITLE_FONT
    ws.cell(row=2, column=2, value="HybridRAG v3 -- Offline-First RAG System")
    ws.cell(row=3, column=1, value="Updated:").font = SUBTITLE_FONT
    ws.cell(row=3, column=2, value=date.today().strftime("%Y-%m-%d"))
    ws.cell(row=4, column=1, value="Revision:").font = SUBTITLE_FONT
    ws.cell(row=4, column=2, value="v5 (full rebuild, justification notes, CVE status)")

    row = 6
    _apply_header(ws, row, ["Category", "Status", "Count (Direct)",
                            "Count (Transitive)", "Est. Size"], BLUE_FILL)
    summaries = [
        ["GREEN", "Approved & Installed", "35", "25", "~200 MB"],
        ["YELLOW", "Applying for Approval", "4 + Ollama + 6 models", "4", "~50 MB + Ollama"],
        ["BLUE", "Recommended (Not Installed)", "4", "1", "~350 MB"],
        ["RED", "Banned (DO NOT SUBMIT)", "11", "--", "N/A"],
        ["RETIRED", "Removed from Stack", "8", "--", "-2.5 GB saved"],
    ]
    fills = [LIGHT_GREEN, LIGHT_YELLOW, LIGHT_BLUE, LIGHT_RED, LIGHT_GRAY]
    for i, (vals, fill) in enumerate(zip(summaries, fills)):
        _apply_data_row(ws, row + 1 + i, vals, fill)

    row = 13
    ws.cell(row=row, column=1, value="Approval Leverage (Already-Approved Anchors):").font = SUBTITLE_FONT
    row += 1
    _apply_header(ws, row, ["If This Is Approved", "Then These Are Already Covered",
                            "", "", ""], GREEN_FILL)
    anchors = [
        ["Python 3.x", "typing_extensions, pip, setuptools (ship with Python)"],
        ["pip", "cryptography, urllib3, requests (pip's own dependencies)"],
        ["Chrome / Edge", "pypdfium2 (same Google PDFium engine as Chrome's PDF viewer)"],
        ["Microsoft Office", "python-docx, python-pptx, openpyxl (just READ Office file formats)"],
        ["Any AI/ML approval", "numpy is the universal foundation for all Python numerical work"],
    ]
    for i, vals in enumerate(anchors):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=vals[0]).font = DATA_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        ws.cell(row=r, column=2, value=vals[1]).font = DATA_FONT
        ws.cell(row=r, column=2).border = THIN_BORDER
        ws.cell(row=r, column=2).alignment = WRAP_ALIGN

    row = 21
    ws.cell(row=row, column=1, value="What to Apply for Next:").font = SUBTITLE_FONT
    row += 1
    _apply_header(ws, row, ["Priority", "Package", "Version", "License", "Action"], YELLOW_FILL)
    priorities = [
        ["P1 (Now)", "openai", "1.109.1", "MIT", "Apply for approval (Azure API client)"],
        ["P1 (Now)", "pytest", "9.0.2", "MIT", "Apply for approval (test framework)"],
        ["P1 (Now)", "psutil", "7.2.2", "BSD-3", "Apply for approval (process monitoring)"],
        ["P1 (Now)", "Ollama", "latest", "MIT", "Apply for approval (offline LLM server)"],
        ["P2 (Scale)", "faiss-cpu", "1.9.0", "MIT", "Apply for approval (vector search)"],
        ["P2 (Scale)", "lancedb", "0.29.2", "Apache 2.0", "Apply for approval (vector DB)"],
        ["P2 (Scale)", "pyarrow", ">=16.0", "Apache 2.0", "Apply for approval (lancedb dep)"],
        ["P3 (Future)", "vllm", "0.10.1", "Apache 2.0", "Apply when workstation arrives"],
    ]
    for i, vals in enumerate(priorities):
        _apply_data_row(ws, row + 1 + i, vals)

    _set_col_widths(ws, [20, 50, 12, 12, 50])

    # -----------------------------------------------------------------------
    # SHEET 2: GREEN -- Approved & Installed
    # -----------------------------------------------------------------------
    ws2 = wb.create_sheet("GREEN - Approved")
    ws2.sheet_properties.tabColor = "228B22"

    ws2.merge_cells("A1:I1")
    ws2.cell(row=1, column=1, value="GREEN -- Approved and Installed (35 direct packages)").font = TITLE_FONT

    _apply_header(ws2, 3, GREEN_HEADERS, GREEN_FILL)
    for i, vals in enumerate(GREEN_DATA):
        fill = LIGHT_GREEN if i % 2 == 0 else WHITE_FILL
        _apply_data_row(ws2, 4 + i, vals, fill)

    # Transitive section
    trans_start = 4 + len(GREEN_DATA) + 2
    ws2.merge_cells(f"A{trans_start}:F{trans_start}")
    ws2.cell(row=trans_start, column=1,
             value="Transitive Dependencies (auto-installed, no separate waiver needed)").font = SUBTITLE_FONT

    _apply_header(ws2, trans_start + 1, GREEN_TRANSITIVE_HEADERS, GREEN_FILL)
    for i, vals in enumerate(GREEN_TRANSITIVE_DATA):
        fill = LIGHT_GREEN if i % 2 == 0 else WHITE_FILL
        _apply_data_row(ws2, trans_start + 2 + i, vals, fill)

    _set_col_widths(ws2, [5, 22, 12, 12, 25, 20, 40, 60, 18])

    # -----------------------------------------------------------------------
    # SHEET 3: YELLOW -- Applying for Approval
    # -----------------------------------------------------------------------
    ws3 = wb.create_sheet("YELLOW - Applying")
    ws3.sheet_properties.tabColor = "DAA520"

    ws3.merge_cells("A1:K1")
    ws3.cell(row=1, column=1,
             value="YELLOW -- Applying for Approval (Currently Installed)").font = TITLE_FONT

    ws3.cell(row=2, column=1, value="Python Packages:").font = SUBTITLE_FONT
    _apply_header(ws3, 3, YELLOW_HEADERS, YELLOW_FILL)
    for i, vals in enumerate(YELLOW_DATA):
        fill = LIGHT_YELLOW if i % 2 == 0 else WHITE_FILL
        _apply_data_row(ws3, 4 + i, vals, fill)

    # Ollama models subsection
    model_start = 4 + len(YELLOW_DATA) + 2
    ws3.cell(row=model_start, column=1,
             value="Ollama AI Models (all installed, run locally on localhost):").font = SUBTITLE_FONT
    _apply_header(ws3, model_start + 1, YELLOW_MODELS_HEADERS, YELLOW_FILL)
    for i, vals in enumerate(YELLOW_MODELS_DATA):
        fill = LIGHT_YELLOW if i % 2 == 0 else WHITE_FILL
        _apply_data_row(ws3, model_start + 2 + i, vals, fill)

    # Pytest transitive deps
    pytest_start = model_start + 2 + len(YELLOW_MODELS_DATA) + 2
    ws3.cell(row=pytest_start, column=1,
             value="Transitive Dependencies (pytest):").font = SUBTITLE_FONT
    _apply_header(ws3, pytest_start + 1,
                  ["#", "Package", "Version", "License", "Pulled In By",
                   "", "", "", "", "", ""], YELLOW_FILL)
    pytest_trans = [
        [1, "iniconfig", "2.3.0", "MIT", "pytest"],
        [2, "pluggy", "1.6.0", "MIT", "pytest"],
        [3, "importlib_metadata", "8.7.1", "Apache 2.0", "pytest"],
        [4, "zipp", "3.23.0", "MIT", "importlib_metadata"],
    ]
    for i, vals in enumerate(pytest_trans):
        _apply_data_row(ws3, pytest_start + 2 + i, vals)

    _set_col_widths(ws3, [5, 22, 14, 12, 25, 20, 50, 65, 40, 50, 18])

    # -----------------------------------------------------------------------
    # SHEET 4: BLUE -- Recommended Future
    # -----------------------------------------------------------------------
    ws4 = wb.create_sheet("BLUE - Recommended")
    ws4.sheet_properties.tabColor = "4169E1"

    ws4.merge_cells("A1:J1")
    ws4.cell(row=1, column=1,
             value="BLUE -- Recommended for Next Phase (Not Yet Installed)").font = TITLE_FONT

    _apply_header(ws4, 3, BLUE_HEADERS, BLUE_FILL)
    for i, vals in enumerate(BLUE_DATA):
        fill = LIGHT_BLUE if i % 2 == 0 else WHITE_FILL
        _apply_data_row(ws4, 4 + i, vals, fill)

    _set_col_widths(ws4, [5, 18, 10, 12, 30, 55, 70, 35, 35, 30])

    # -----------------------------------------------------------------------
    # SHEET 5: RED -- Banned
    # -----------------------------------------------------------------------
    ws5 = wb.create_sheet("RED - Banned")
    ws5.sheet_properties.tabColor = "CC3333"

    ws5.merge_cells("A1:G1")
    ws5.cell(row=1, column=1,
             value="RED -- Banned Software (DO NOT SUBMIT)").font = TITLE_FONT

    _apply_header(ws5, 3, RED_HEADERS, RED_FILL)
    for i, vals in enumerate(RED_DATA):
        fill = LIGHT_RED if i % 2 == 0 else WHITE_FILL
        _apply_data_row(ws5, 4 + i, vals, fill)

    _set_col_widths(ws5, [5, 25, 20, 18, 15, 60, 45])

    # -----------------------------------------------------------------------
    # SHEET 6: RETIRED
    # -----------------------------------------------------------------------
    ws6 = wb.create_sheet("RETIRED")
    ws6.sheet_properties.tabColor = "708090"

    ws6.merge_cells("A1:G1")
    ws6.cell(row=1, column=1,
             value="RETIRED -- Removed from Stack (2026-02-24)").font = TITLE_FONT
    ws6.cell(row=2, column=1,
             value="Impact: ~2.5 GB removed from virtual environment. No functionality lost.").font = DATA_FONT

    _apply_header(ws6, 4, RETIRED_HEADERS, GRAY_FILL)
    for i, vals in enumerate(RETIRED_DATA):
        fill = LIGHT_GRAY if i % 2 == 0 else WHITE_FILL
        _apply_data_row(ws6, 5 + i, vals, fill)

    _set_col_widths(ws6, [5, 25, 14, 12, 50, 40, 12])

    return wb


def main():
    output_path = os.path.join(PROJECT_ROOT, "docs", "05_security",
                               "waiver_cheat_sheet_v5.xlsx")
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = create_workbook()
    wb.save(output_path)
    print("[OK] Generated: %s" % output_path)
    print("     Sheets: Summary, GREEN, YELLOW, BLUE, RED, RETIRED")
    print("     35 GREEN direct + 25 transitive")
    print("     4 YELLOW packages + Ollama + 6 models")
    print("     4 BLUE recommendations")
    print("     11 RED banned entries")
    print("     8 RETIRED packages")


if __name__ == "__main__":
    main()
